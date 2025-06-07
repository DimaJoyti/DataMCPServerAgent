"""
Excel file parser for .xlsx and .xls files.
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .base_parser import BaseParser, ParsedDocument
from ..metadata.models import DocumentMetadata, DocumentType

class ExcelParser(BaseParser):
    """Parser for Excel files (.xlsx, .xls)."""

    def __init__(self):
        """Initialize Excel parser."""
        super().__init__()
        self.logger = logging.getLogger(self.__class__.__name__)

        if not HAS_PANDAS:
            raise ImportError("pandas is required for Excel parsing. Install with: pip install pandas")

        if not HAS_OPENPYXL:
            self.logger.warning("openpyxl not available. Some Excel features may not work. Install with: pip install openpyxl")

    def can_parse(self, file_path: Path) -> bool:
        """Check if this parser can handle the file."""
        return file_path.suffix.lower() in ['.xlsx', '.xls']

    def parse(self, file_path: Path, **kwargs) -> ParsedDocument:
        """
        Parse Excel file.

        Args:
            file_path: Path to Excel file
            **kwargs: Additional parsing options
                - sheet_names: List of sheet names to parse (default: all)
                - include_headers: Whether to include column headers (default: True)
                - max_rows: Maximum number of rows to parse per sheet (default: None)
                - format_as_table: Whether to format as table (default: True)

        Returns:
            ParsedDocument: Parsed document
        """
        try:
            # Parse options
            sheet_names = kwargs.get('sheet_names', None)  # None means all sheets
            include_headers = kwargs.get('include_headers', True)
            max_rows = kwargs.get('max_rows', None)
            format_as_table = kwargs.get('format_as_table', True)

            # Read Excel file
            if sheet_names:
                # Read specific sheets
                excel_data = pd.read_excel(
                    file_path,
                    sheet_name=sheet_names,
                    nrows=max_rows,
                    header=0 if include_headers else None
                )
            else:
                # Read all sheets
                excel_data = pd.read_excel(
                    file_path,
                    sheet_name=None,  # Read all sheets
                    nrows=max_rows,
                    header=0 if include_headers else None
                )

            # Handle single sheet case
            if isinstance(excel_data, pd.DataFrame):
                excel_data = {'Sheet1': excel_data}

            # Extract text content
            text_content = []
            metadata = {}

            # Process each sheet
            for sheet_name, df in excel_data.items():
                if df.empty:
                    continue

                sheet_text = self._dataframe_to_text(df, sheet_name, format_as_table, include_headers)
                text_content.append(sheet_text)

                # Collect sheet metadata
                metadata[f'sheet_{sheet_name}'] = {
                    'rows': len(df),
                    'columns': len(df.columns),
                    'column_names': list(df.columns) if include_headers else None,
                    'has_data': not df.empty
                }

            # Combine all text
            full_text = '\n\n'.join(text_content)

            # Create document metadata
            doc_metadata = DocumentMetadata(
                title=file_path.stem,
                document_type=DocumentType.SPREADSHEET,
                file_path=str(file_path),
                file_size=file_path.stat().st_size,
                language="unknown",  # Excel doesn't have inherent language
                page_count=len(excel_data),  # Number of sheets
                word_count=len(full_text.split()),
                character_count=len(full_text),
                custom_metadata={
                    'total_sheets': len(excel_data),
                    'sheet_names': list(excel_data.keys()),
                    'sheets_metadata': metadata,
                    'parser': 'ExcelParser'
                }
            )

            return ParsedDocument(
                text=full_text,
                metadata=doc_metadata,
                raw_content=excel_data,  # Store original DataFrames
                processing_time=0.0  # Will be set by caller
            )

        except Exception as e:
            self.logger.error(f"Failed to parse Excel file {file_path}: {e}")
            raise

    def _dataframe_to_text(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        format_as_table: bool,
        include_headers: bool
    ) -> str:
        """Convert DataFrame to text representation."""
        if df.empty:
            return f"Sheet: {sheet_name}\n(Empty sheet)"

        text_parts = [f"Sheet: {sheet_name}"]

        if format_as_table:
            # Format as a readable table
            if include_headers:
                # Add column headers
                headers = " | ".join(str(col) for col in df.columns)
                text_parts.append(headers)
                text_parts.append("-" * len(headers))

            # Add data rows
            for _, row in df.iterrows():
                row_text = " | ".join(str(value) if pd.notna(value) else "" for value in row)
                text_parts.append(row_text)
        else:
            # Format as structured text
            for idx, row in df.iterrows():
                if include_headers:
                    row_parts = []
                    for col, value in row.items():
                        if pd.notna(value):
                            row_parts.append(f"{col}: {value}")
                    if row_parts:
                        text_parts.append(f"Row {idx + 1}: " + ", ".join(row_parts))
                else:
                    row_values = [str(value) for value in row if pd.notna(value)]
                    if row_values:
                        text_parts.append(f"Row {idx + 1}: " + ", ".join(row_values))

        return "\n".join(text_parts)

    def extract_metadata(self, file_path: Path) -> Dict:
        """Extract metadata from Excel file."""
        try:
            # Try to get workbook metadata using openpyxl
            if HAS_OPENPYXL and file_path.suffix.lower() == '.xlsx':
                return self._extract_openpyxl_metadata(file_path)
            else:
                # Fallback to pandas-based metadata
                return self._extract_pandas_metadata(file_path)

        except Exception as e:
            self.logger.warning(f"Failed to extract Excel metadata: {e}")
            return {}

    def _extract_openpyxl_metadata(self, file_path: Path) -> Dict:
        """Extract metadata using openpyxl."""
        from openpyxl import load_workbook

        metadata = {}

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            # Basic workbook info
            metadata.update({
                'sheet_names': workbook.sheetnames,
                'total_sheets': len(workbook.sheetnames),
                'active_sheet': workbook.active.title if workbook.active else None
            })

            # Document properties
            props = workbook.properties
            if props:
                metadata.update({
                    'title': props.title,
                    'creator': props.creator,
                    'description': props.description,
                    'subject': props.subject,
                    'keywords': props.keywords,
                    'category': props.category,
                    'created': props.created.isoformat() if props.created else None,
                    'modified': props.modified.isoformat() if props.modified else None,
                    'last_modified_by': props.lastModifiedBy,
                    'revision': props.revision,
                    'version': props.version
                })

            # Sheet-specific metadata
            sheets_info = {}
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheets_info[sheet_name] = {
                    'max_row': sheet.max_row,
                    'max_column': sheet.max_column,
                    'title': sheet.title,
                    'sheet_state': sheet.sheet_state,
                    'sheet_type': str(type(sheet).__name__)
                }

            metadata['sheets_info'] = sheets_info

            workbook.close()

        except Exception as e:
            self.logger.warning(f"Failed to extract openpyxl metadata: {e}")

        return metadata

    def _extract_pandas_metadata(self, file_path: Path) -> Dict:
        """Extract metadata using pandas."""
        metadata = {}

        try:
            # Get basic info about all sheets
            excel_file = pd.ExcelFile(file_path)

            metadata.update({
                'sheet_names': excel_file.sheet_names,
                'total_sheets': len(excel_file.sheet_names),
                'engine': excel_file.engine if hasattr(excel_file, 'engine') else 'unknown'
            })

            # Get info for each sheet
            sheets_info = {}
            for sheet_name in excel_file.sheet_names:
                try:
                    # Read just the first few rows to get structure info
                    df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=0)
                    sheets_info[sheet_name] = {
                        'columns': list(df.columns),
                        'column_count': len(df.columns),
                        'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()}
                    }
                except Exception as e:
                    self.logger.warning(f"Failed to get info for sheet {sheet_name}: {e}")
                    sheets_info[sheet_name] = {'error': str(e)}

            metadata['sheets_info'] = sheets_info

        except Exception as e:
            self.logger.warning(f"Failed to extract pandas metadata: {e}")

        return metadata

    def get_supported_extensions(self) -> List[str]:
        """Get list of supported file extensions."""
        return ['.xlsx', '.xls']

    def get_parser_info(self) -> Dict:
        """Get information about this parser."""
        return {
            'name': 'ExcelParser',
            'description': 'Parser for Microsoft Excel files',
            'supported_extensions': self.get_supported_extensions(),
            'features': [
                'Multiple sheet support',
                'Table formatting',
                'Metadata extraction',
                'Column header preservation',
                'Configurable row limits'
            ],
            'dependencies': {
                'pandas': HAS_PANDAS,
                'openpyxl': HAS_OPENPYXL
            }
        }
